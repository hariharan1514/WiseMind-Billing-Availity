import os.path
from dataclasses import replace
from datetime import datetime
import pandas as pd
import sys

from pymsgbox import confirm
from selenium.common import NoSuchElementException
from selenium.webdriver import ActionChains

parent_folder_path = r"Z:\Wisemind\Charge Entry -Billing\Billing Dates"
path_temp_date = datetime.today().strftime('%m%d%Y')
bcbs_file_path = (parent_folder_path + "\\" +path_temp_date[4:] + "\\" +datetime.today().strftime("%m %b'%Y") +
                   "\\" +path_temp_date + "\\" +f"BCBS scrubbed file - {path_temp_date}.xlsx")

if not os.path.exists(bcbs_file_path):
    print("Run the 'WiseMind_Billing_Phase3.py' script first, then execute the Availity billing script.")
    sys.exit()
else:
    bcbs_billing_df = pd.read_excel(bcbs_file_path, sheet_name=0)
    print(f"Availity Data Row Count: {len(bcbs_billing_df)}")
    ### Check if the Straightforward Billing file contains at least one data row. ###
    if bcbs_billing_df.shape[0] == 0:
        print(
            "No Availity billing cases detected in today's run. Script completed successfully with no records to process. \n\nExiting gracefully as per expected behavior.")
        sys.exit()
    else:
        print("Good to go !!!")
        import time
        import openpyxl
        from openpyxl import Workbook, load_workbook
        from selenium import webdriver
        from selenium.webdriver.ie.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.wait import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        import re

        # Read Configuration Sheet with openpyxl & Pandas module
        config_sheet_path = r"Z:\Wisemind\Charge Entry -Billing\Automation Config File\ConfigSheet.xlsx"
        # Get the Username & Password
        master_workbook = load_workbook(config_sheet_path, data_only=True)
        password_sheet = master_workbook[master_workbook.sheetnames[0]]
        username = password_sheet['B5'].value
        password = password_sheet['B6'].value
        claim_encounter_url = password_sheet['B10'].value

        availitypayor_df = pd.read_excel(config_sheet_path, sheet_name=4)
        availitypayor_staffmember_dict = availitypayor_df.set_index('Staff Members')[
            ['Availity Rendering Provider', 'Availity Billing Provider']].to_dict(orient='index')

        # Initiate the Chrome instance
        chrome_option = webdriver.ChromeOptions()
        chrome_option.add_experimental_option('detach', True)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_option)
        driver.maximize_window()
        actions = ActionChains(driver)

        driver.get("https://apps.availity.com/web/onboarding/availity-fr-ui/#/login")  # Launching the router

        # send the username & password to the represented field
        try:
            username_element = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.XPATH, "//input[@id='userId']")))
            username_element.send_keys(username)
        except:
            print("Login issue, Please login after some time.!")
            sys.exit()

        password_element = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@id='password']")))
        password_element.send_keys(password)

        login_button_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Sign In']")))
        login_button_element.click()

        code = input("Please enter the code manually and hit enter here:")
        time.sleep(7)
        popupCheck = input(
            "Please clear all pop content then hit enter if not pop up content then go ahead and hit enter:")
        time.sleep(7)

        try:
            portalerror_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='card-title']")))
            print("Oops! Something went wrong. Please try logging in again later.")
            driver.quit()
            sys.exit()
        except:
            pass

        driver.get(claim_encounter_url)
        time.sleep(3)

        ## Insurance Company/Benefit Plan Information ###

        billing_frame_element = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//iframe[@id='newBodyFrame']")))
        driver.switch_to.frame(billing_frame_element)

        organization_element = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@name='organization']")))
        organization_element.click()
        organization_element.send_keys("Wise Mind Psychological Services, P.L.L.C.")
        time.sleep(2)
        driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
        time.sleep(1)
        driver.switch_to.active_element.send_keys(Keys.ENTER)

        claim_type_element = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@name='transactionType']")))
        claim_type_element.click()
        claim_type_element.clear()
        claim_type_element.send_keys("Professional Claim")
        time.sleep(2)
        driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
        time.sleep(1)
        driver.switch_to.active_element.send_keys(Keys.ENTER)

        payor_type_element = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@name='payer']")))
        payor_type_element.click()
        payor_type_element.clear()
        payor_type_element.send_keys("ANTHEM BCBS NY")
        time.sleep(2)
        driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
        time.sleep(1)
        driver.switch_to.active_element.send_keys(Keys.ENTER)

        availity_billing_wb = load_workbook(bcbs_file_path, data_only=True)
        availity_billing_ws = availity_billing_wb.active

        ### Add Needed columns
        needed_clms = ["Transaction Number", "Status"]
        availity_billing_wb_headers = [cell.value for cell in availity_billing_ws[1]]
        next_col_index = len(availity_billing_wb_headers) + 1
        for col in needed_clms:
            if col not in availity_billing_wb_headers:
                availity_billing_ws.cell(row=1, column=next_col_index, value=col)
                next_col_index += 1
        availity_billing_wb.save(bcbs_file_path)

        ### Build a dictionary mapping header names to column indices
        columns = {}
        for col in range(1, availity_billing_ws.max_column + 1):
            col_name = availity_billing_ws.cell(row=1, column=col).value
            if col_name:
                columns[col_name.strip()] = col

        for row in range(2, availity_billing_ws.max_row + 1):

            status = availity_billing_ws.cell(row=row, column=columns['Status']).value

            if status == "Yes":
                continue

            client_name = availity_billing_ws.cell(row=row, column=columns['Client Name']).value
            client_id_number = availity_billing_ws.cell(row=row, column=columns['Client ID Number']).value
            service_date = availity_billing_ws.cell(row=row, column=columns['Date/Time']).value
            cpt_code = availity_billing_ws.cell(row=row, column=columns['Service Type']).value
            staffmember = availity_billing_ws.cell(row=row, column=columns['Staff Member(s)']).value
            payor_name = availity_billing_ws.cell(row=row, column=columns['Payor Name']).value
            first_name = availity_billing_ws.cell(row=row, column=columns['First Name']).value
            last_name = availity_billing_ws.cell(row=row, column=columns['Last Name']).value
            dob = availity_billing_ws.cell(row=row, column=columns['DOB']).value
            gender = availity_billing_ws.cell(row=row, column=columns['Gender']).value
            street = availity_billing_ws.cell(row=row, column=columns['Street']).value
            city = availity_billing_ws.cell(row=row, column=columns['City']).value
            state = availity_billing_ws.cell(row=row, column=columns['State']).value
            zip_code = availity_billing_ws.cell(row=row, column=columns['ZIP Code']).value
            claim_invoice_number = availity_billing_ws.cell(row=row, column=columns['Claim Invoice Number']).value
            insurance_number = availity_billing_ws.cell(row=row, column=columns['Insurance Number']).value
            place_of_service = availity_billing_ws.cell(row=row, column=columns['Place of Service']).value
            dx_code = availity_billing_ws.cell(row=row, column=columns['DX Code']).value
            charge_amount = availity_billing_ws.cell(row=row, column=columns['Charge Amount']).value
            quantity = availity_billing_ws.cell(row=row, column=columns['Quantity']).value

            ### PATIENT INFORMATION ###

            lastname_element = WebDriverWait(driver, 120).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.lastName']")))
            lastname_element.send_keys(last_name)
            time.sleep(1)

            firstname_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.firstName']")))
            firstname_element.send_keys(first_name)

            dob_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.birthDate']")))
            dob_element.send_keys(dob)

            ### Gender Need to bee add logic
            if gender == "Male" :

                gender_element = WebDriverWait(driver, 60).until(
                    EC.visibility_of_element_located((By.XPATH, "//input[@name='patient.genderCode']")))
                gender_element.click()
                gender_element.clear()
                gender_element.send_keys(gender)
                time.sleep(2)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                time.sleep(1)
                driver.switch_to.active_element.send_keys(Keys.ENTER)

            else:
                gender_element = WebDriverWait(driver, 60).until(
                    EC.visibility_of_element_located((By.XPATH, "//input[@name='patient.genderCode']")))
                gender_element.click()
                gender_element.clear()
                gender_element.send_keys(gender)
                time.sleep(2)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                time.sleep(1)
                driver.switch_to.active_element.send_keys(Keys.ENTER)

            adress_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.addressLine1']")))
            adress_element.send_keys(street)

            city_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.city']")))
            city_element.send_keys(city)

            state_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.stateCode']")))
            state_element.send_keys("New York")
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            zipcode_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='patient.zipCode']")))
            zipcode_element.send_keys(zip_code)

            ### SUBSCRIBER INFORMATION ###

            insuranceid_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='subscriber.memberId']")))
            insuranceid_element.send_keys(insurance_number)

            authorized_plan_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.benefitsAssignmentCertification']")))
            authorized_plan_element.send_keys("Y -")
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            ### BILLING PROVIDER INFORMATION ###

            if staffmember in availitypayor_staffmember_dict:
                rendering_provider = availitypayor_staffmember_dict[staffmember]['Availity Rendering Provider']
                billing_provider = availitypayor_staffmember_dict[staffmember]['Availity Billing Provider']
            else:
                print(
                    f"The Staff name ({staffmember}) is missing in the Staff Member table masters. Kindly check and re run the script")
                sys.exit()

            # rendering_provider_extracted = re.sub(r"\s*\(NPI:.*\)", "", rendering_provider).strip()
            # billing_provider_extracted = re.sub(r"\s*\(NPI:.*\)", "", billing_provider).strip()

            rendering_provider_extracted = rendering_provider.split(",")[0].strip()
            billing_provider_extracted = billing_provider.split(",")[0].strip()

            select_billing_provider_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "(//div[@class='MuiBox-root css-5749k3'])[3]//input")))
            select_billing_provider_element.send_keys(billing_provider_extracted)
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            time.sleep(3)
            driver.switch_to.active_element.send_keys(Keys.TAB)
            time.sleep(1)

            # select_address_element = WebDriverWait(driver, 60).until(
            #     EC.element_to_be_clickable((By.XPATH, "//input[@id=':r14:']")))
            driver.switch_to.active_element.send_keys("77 North Centre Ave")
            time.sleep(2)
            address = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), '77 North Centre Ave') or contains(text(), '77 N CENTRE AVE STE 310')]")))
            address.click()
            # driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            # time.sleep(1)
            # driver.switch_to.active_element.send_keys(Keys.ENTER)

            try:
                select_rendering_provider_element = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "(//div[@class='MuiBox-root css-1fyjx9k'][1]//input)[1]")))
            except:
                ### Add Rendering Provider ###
                add_renderingprovider_element = WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Add Rendering Provider']")))
                add_renderingprovider_element.click()

            ### RENDERING PROVIDER ###
            select_rendering_provider_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "(//div[@class='MuiBox-root css-1fyjx9k'][1]//input)[1]")))
            select_rendering_provider_element.send_keys(rendering_provider_extracted)
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            ### CLAIM INFORMATION ###

            patient_contolnumber_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.controlNumber']")))
            patient_contolnumber_element.click()
            patient_contolnumber_element.send_keys(claim_invoice_number)

            place_of_service_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.placeOfServiceCode']")))
            place_of_service_element.click()
            place_of_service_element.send_keys(place_of_service)
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            frequencytype_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.frequencyTypeCode']")))
            frequencytype_element.click()
            frequencytype_element.send_keys("1")
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            provider_accepts_assignment_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.providerAcceptAssignmentCode']")))
            provider_accepts_assignment_element.click()
            provider_accepts_assignment_element.send_keys("A -")
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            release_of_information_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.informationReleaseCode']")))
            release_of_information_element.click()
            release_of_information_element.send_keys("Y -")
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            provider_signature_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.providerSignatureOnFile']")))
            provider_signature_element.click()
            provider_signature_element.send_keys("Yes")
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            ### DIAGNOSIS CODES ###

            dx_code_list = dx_code.split(",")

            if len(dx_code_list) > 1 :
                for code in range(0, len(dx_code_list), +1):
                    if code == 0:
                        dxcode_box_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.diagnoses.0.code']")))
                        dxcode_box_element.click()
                        dxcode_box_element.send_keys(dx_code_list[code])
                        time.sleep(2)
                        driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                        time.sleep(1)
                        driver.switch_to.active_element.send_keys(Keys.ENTER)
                    else:
                        add_button_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Add button']")))
                        add_button_element.click()

                        dxcode_box_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable((By.XPATH, f"//input[@name='claimInformation.diagnoses.{code}.code']")))
                        dxcode_box_element.click()
                        dxcode_box_element.send_keys(dx_code_list[code])
                        time.sleep(2)
                        driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                        time.sleep(1)
                        driver.switch_to.active_element.send_keys(Keys.ENTER)
            else:
                dxcode_box_element = WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.diagnoses.0.code']")))
                dxcode_box_element.click()
                dxcode_box_element.send_keys(dx_code)
                time.sleep(2)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                time.sleep(1)
                driver.switch_to.active_element.send_keys(Keys.ENTER)

            ### LINES ###

            extracted_service_date = service_date.split(",")[0]

            service_from_date_element = WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.fromDate']")))
            service_from_date_element.click()
            service_from_date_element.send_keys(extracted_service_date)

            service_to_date_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.toDate']")))
            service_to_date_element.click()
            service_to_date_element.send_keys(extracted_service_date)

            placeofservice_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.placeOfServiceCode']")))
            placeofservice_element.click()
            placeofservice_element.send_keys(place_of_service)
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            extracted_procedure_code = cpt_code.split(":")[0]

            procedurecode_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.procedureCode']")))
            procedurecode_element.click()
            procedurecode_element.send_keys(extracted_procedure_code)
            time.sleep(2)
            driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)
            driver.switch_to.active_element.send_keys(Keys.ENTER)

            ## DX Pointer logig need to be work ##

            dxcode_count = (len(dx_code_list))

            for code_index in range(0, len(dx_code_list), +1):
                dxpointer_element = WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.XPATH,f"//input[@name='claimInformation.serviceLines.0.diagnosisCodePointer{code_index+1}']")))
                dxpointer_element.click()
                dxpointer_element.send_keys(dx_code_list[code_index])
                time.sleep(2)
                driver.switch_to.active_element.send_keys(Keys.ARROW_DOWN)
                time.sleep(1)
                driver.switch_to.active_element.send_keys(Keys.ENTER)

            charge_amount_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.amount']")))
            charge_amount_element.click()
            charge_amount_element.send_keys(charge_amount)

            quantity_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='claimInformation.serviceLines.0.quantity']")))
            quantity_element.click()
            quantity_element.send_keys(quantity)

            ### Continue & Submitt & Transaction ID logic need to done. ###

            continue_button_element = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
            continue_button_element.click()

            submit_button = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Submit']")))
            submit_button.click()

            transaction_id_element = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.XPATH, "//p[contains(@class, 'MuiTypography-root') and contains(text(), 'Transaction ID')]/following-sibling::p")))
            transaction_id = transaction_id_element.text.strip()

            availity_billing_ws.cell(row=row, column=columns['Transaction Number']).value = transaction_id
            availity_billing_ws.cell(row=row, column=columns['Status']).value = "Yes"
            availity_billing_wb.save(bcbs_file_path)
            print(f"✅ The Claim : {client_name} has been billed." )

            new_claim = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='New Claim']")))
            new_claim.click()
            time.sleep(3)

        driver.switch_to.default_content()
        logout = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Logout']")))
        logout.click()
        time.sleep(10)
        driver.quit()









































































