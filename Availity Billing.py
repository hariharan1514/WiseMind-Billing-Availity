import os.path
from dataclasses import replace
from datetime import datetime
import pandas as pd
import sys

from pymsgbox import confirm
from selenium.common import NoSuchElementException
from selenium.webdriver import ActionChains

parent_folder_path = r"Z:\Wisemind\Charge Entry -Billing\Billing Dates"
path_temp_date = datetime.today().strftime('%m%d%Y')
bcbs_file_path = (parent_folder_path + "\\" +path_temp_date[4:] + "\\" +datetime.today().strftime("%m %b'%Y") +
                   "\\" +path_temp_date + "\\" +f"BCBS scrubbed file - {path_temp_date}.xlsx")

if not os.path.exists(bcbs_file_path):
    print("Run the 'WiseMind_Billing_Phase3.py' script first, then execute the Availity billing script.")
    sys.exit()
else:
    bcbs_billing_df = pd.read_excel(bcbs_file_path, sheet_name=0)
    print(f"Availity Data Row Count: {len(bcbs_billing_df)}")
    ### Check if the Straightforward Billing file contains at least one data row. ###
    if bcbs_billing_df.shape[0] == 0:
        print(
            "No Availity billing cases detected in today's run. Script completed successfully with no records to process. \n\nExiting gracefully as per expected behavior.")
        sys.exit()
    else:
        print("Good to go !!!")
        import time
        from openpyxl import Workbook, load_workbook
        from selenium import webdriver
        from selenium.webdriver.ie.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.wait import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        import re

        # Read Configuration Sheet with openpyxl & Pandas module
        config_sheet_path = r"Z:\Wisemind\Charge Entry -Billing\Automation Config File\ConfigSheet.xlsx"
        # Get the Username & Password
        master_workbook = load_workbook(config_sheet_path, data_only=True)
        password_sheet = master_workbook[master_workbook.sheetnames[0]]
        username = password_sheet['B5'].value
        password = password_sheet['B6'].value

        availitypayor_df = pd.read_excel(config_sheet_path, sheet_name=4)
        availitypayor_staffmember_dict = availitypayor_df.set_index('Staff Members')[
            ['Availity Rendering Provider', 'Availity Billing Provider']].to_dict(orient='index')

        # Initiate the Chrome instance
        chrome_option = webdriver.ChromeOptions()
        chrome_option.add_experimental_option('detach', True)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_option)
        driver.maximize_window()
        actions = ActionChains(driver)

        driver.get("https://apps.availity.com/web/onboarding")  # Launching the router

        # send the username & password to the represented field
        try:
            username_element = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.XPATH, "//input[@id='userId']")))
            username_element.send_keys(username)
        except:
            print("Login issue, Please login after some time.!")
            sys.exit()

        password_element = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@id='password']")))
        password_element.send_keys(password)

        login_button_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Sign In']")))
        login_button_element.click()

        # 2 Step Authentication
        textme_button_element = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "(//span//span[@class='css-1qiat4j'])[1]")))
        print("finded")
        driver.execute_script("arguments[0].click();", textme_button_element)
        # textme_button_element.click()

        continue_button_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Continue']")))
        continue_button_element.click()

        autentication_code = int(input("Kindly Enter the OTP Code :"))
        time.sleep(10)

        code_box_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@id='code']")))
        code_box_element.send_keys(autentication_code)

        autentication_continue_button_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Continue']")))
        autentication_continue_button_element.click()

        try:
            confirmation_continue_button_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Continue']")))
            confirmation_continue_button_element.click()
        except:
            pass

        try:
            portalerror_element = WebDriverWait(driver, 90).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='card-title']")))
            print("Oops! Something went wrong. Please try logging in again later.")
            driver.quit()
            sys.exit()
        except:
            pass