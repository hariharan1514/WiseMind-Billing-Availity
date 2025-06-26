import os.path
from datetime import datetime
import pandas as pd
import sys

from pymsgbox import confirm
from selenium.common import NoSuchElementException
from selenium.webdriver import ActionChains

parent_folder_path = r"Z:\Wisemind\Charge Entry -Billing\Billing Dates"
path_temp_date = datetime.today().strftime('%m%d%Y')
billing_file_path = (parent_folder_path + "\\" +path_temp_date[4:] + "\\" +datetime.today().strftime("%m %b'%Y") +
                   "\\" +path_temp_date + "\\" +f"Straightforward Billing Case - {path_temp_date}.xlsx")
bcbs_file_path = (parent_folder_path + "\\" +path_temp_date[4:] + "\\" +datetime.today().strftime("%m %b'%Y") +
                   "\\" +path_temp_date + "\\" +f"BCBS scrubbed file - {path_temp_date}.xlsx")

### Check whether the Straightforward Billing file is present or not.
if not os.path.exists(billing_file_path):
    print("Run the 'WiseMind_Billing_Phase2.py' script first, then execute the Starightforward cases billing script.")
    sys.exit()
else:
    sf_billing_df = pd.read_excel(billing_file_path, sheet_name=0)
    print(f"Staright Forward Data Row Count: {len(sf_billing_df)}")
    ### Check if the Straightforward Billing file contains at least one data row. ###
    if sf_billing_df.shape[0] == 0:
        print("No straightforward billing cases detected in today's run. Script completed successfully with no records to process. \n\nExiting gracefully as per expected behavior.")
        sys.exit()
    else:
        import time
        from openpyxl import load_workbook
        from selenium import webdriver
        from selenium.webdriver.ie.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.wait import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        import re

        # Read Configuration Sheet with openpyxl & Pandas module
        config_sheet_path = r"Z:\Wisemind\Charge Entry -Billing\Automation Config File\ConfigSheet.xlsx"
        # Get the Username & Password
        master_workbook = load_workbook(config_sheet_path, data_only=True)
        password_sheet = master_workbook[master_workbook.sheetnames[0]]
        username = password_sheet['B1'].value
        password = password_sheet['B2'].value

        # Get Staff Member wise Payors
        staff_mem_df = pd.read_excel(config_sheet_path, sheet_name= 1)
        staff_member_payors = staff_mem_df['Staff Member payors'].dropna().tolist()

        # Set Payor wise Billing and Rendering provider & Set Staff Member wise Billing and Rendering provider
        payorwise_provider_df = pd.read_excel(config_sheet_path, sheet_name= 2)
        staffmemberwise_provider_df = pd.read_excel(config_sheet_path, sheet_name= 3)
        availitypayor_df = pd.read_excel(config_sheet_path, sheet_name=4)
        # Convert Provider Table to Provider Dictionary
        payorwise_providerDf_dict = payorwise_provider_df.set_index('Payer')[['Rendering Provider','Billing Provider']].to_dict(orient='index')
        staffmemberwise_providerDf_dict = staffmemberwise_provider_df.set_index('Staff Members')[['Rendering Provider', 'Billing Provider']].to_dict(orient='index')
        availitypayor_staffmember_dict = availitypayor_df.set_index('Staff Members')[['Rendering Provider', 'Billing Provider']].to_dict(orient='index')
        print(availitypayor_staffmember_dict)
        #Availity Payor
        availitypayor_df = pd.read_excel(config_sheet_path, sheet_name=1)
        availitypayor = availitypayor_df['Availity Payors'].dropna().tolist()
        print(availitypayor)

        # Initiate the Chrome instance
        chrome_option = webdriver.ChromeOptions()
        chrome_option.add_experimental_option('detach', True)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_option)
        driver.maximize_window()
        actions = ActionChains(driver)

        driver.get("https://app.theranest.com/login") # Launching the router

        # send the username & password to the represented field
        try:
            username_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='Email']")))
            username_element.send_keys(username)
        except:
            print("Login issue, Please login after some time.!")
            sys.exit()

        password_element = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='Password']")))
        password_element.send_keys(password)

        login_button_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//button[normalize-space()='Log In']")))
        login_button_element.click()

        try:
            mainpage_element = WebDriverWait(driver,90).until(EC.visibility_of_all_elements_located((By.XPATH,"//div[@role='group']")))
        except:
            login_button_element = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Log In']")))
            print("Please try logging in again after some time. The site is currently experiencing login issues.")
            driver.close()
            sys.exit()

        sf_billing_wb = load_workbook(billing_file_path, data_only=True)
        scrubbing_sheet = sf_billing_wb.active

        ### Add Exception Column if not present
        exception_col_name = ["Exceptions","Active/Archived"]
        headers = [cell.value for cell in scrubbing_sheet[1]]
        next_col_index = len(headers) + 1
        for col in exception_col_name:
            if col not in headers:
                scrubbing_sheet.cell(row=1,column=next_col_index,value=col)
                next_col_index += 1
        sf_billing_wb.save(billing_file_path)

        ### Build a dictionary mapping header names to column indices
        data_columns = {}
        for col in range(1, scrubbing_sheet.max_column + 1):
            col_name = scrubbing_sheet.cell(row=1, column=col).value
            if col_name:
                data_columns[col_name.strip()] = col
        print(data_columns)
        for row in range(2, scrubbing_sheet.max_row + 1):

            is_billed = scrubbing_sheet.cell(row=row, column=data_columns['Is Billed']).value
            exception = scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value
            if is_billed == "Yes" or exception != None:
                continue

            client_name = scrubbing_sheet.cell(row=row, column=data_columns['Client Name']).value
            client_id_number = scrubbing_sheet.cell(row=row, column=data_columns['Client ID Number']).value
            dos_date = scrubbing_sheet.cell(row=row, column=data_columns['Date/Time']).value
            staff_member = scrubbing_sheet.cell(row=row, column=data_columns['Staff Member(s)']).value
            payor_name = scrubbing_sheet.cell(row=row, column=data_columns['Payor Name']).value
            claimTab_status = scrubbing_sheet.cell(row=row, column=data_columns['Active/Archived']).value
            status = scrubbing_sheet.cell(row=row, column=data_columns['Status']).value
            service_type = scrubbing_sheet.cell(row=row, column=data_columns['Service Type']).value

            print(f"Patient Name: {client_name}")
            partial_name = client_name.split()[0]
            driver.get("https://wisemind71.theranest.com/clients")
            time.sleep(3)

            tabs = ["Active","Archived"]
            name_to_search = [client_name,partial_name]

            found_data = False
            active_tab = False
            archived_tab =False
            availity_payor = False
            insurance_id_check = False

            ### Check for Availity Payor
            if payor_name in availity_payor:
                availity_payor = True
            else:
                availity_payor = False

            for tab in tabs:
                for name in name_to_search:

                    for attempt in range(3):
                        try:
                            tab_element = WebDriverWait(driver,120).until(
                                EC.presence_of_element_located((By.XPATH,f"(//span[contains(text(), '{tab}')])[1]")))
                            time.sleep(1)
                            tab_element.click()
                            time.sleep(2)
                            break
                        except StaleElementReferenceException:
                            continue

                    search_bar_element = WebDriverWait(driver, 120).until(
                        EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
                    search_bar_element.clear()
                    search_bar_element.send_keys(name)
                    time.sleep(2)
                    #Check if data exists
                    try:
                        # driver.find_element(By.XPATH,"//div[text()='No data available']")
                        no_data_element = WebDriverWait(driver,5).until(EC.visibility_of_element_located((By.XPATH,"//div[text()='No data available']")))
                        search_bar_element.clear()

                    except:
                        found_data = True
                        clientname_table_element = WebDriverWait(driver,120).until(EC.visibility_of_all_elements_located((By.XPATH,"//table[contains(@class, 'k-grid-table')]//tr")))
                        for tbl_row in range(1,len(clientname_table_element) +1):
                            time.sleep(2)
                            if tab == 'Active':
                                client_name_element =WebDriverWait(driver,120).until(EC.element_to_be_clickable((By.XPATH,f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                client_name_element.click()

                                ### Check Client ID found or not
                                client_id_element_number = None
                                try:
                                    client_id_element = WebDriverWait(driver,30).until(EC.visibility_of_element_located((By.XPATH,"//div[@data-aqa='ClientID']//div[2]")))
                                    client_id_element_number = client_id_element.text
                                    # print(f"Captured client ID: {client_id_element_number}")
                                except:
                                    driver.back()
                                    found_data = False

                                if client_id_number == client_id_element_number:
                                    found_data = True
                                    active_tab = True

                                    claim_url = driver.current_url
                                    firstname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='FirstName']"))).get_attribute("value")
                                    middlename = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='MiddleName']"))).get_attribute("value")
                                    lastname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='LastName']"))).get_attribute("value")



                                    portal_client_name = f"{firstname} {middlename} {lastname}".strip()
                                    portal_client_name = " ".join(portal_client_name.split())
                                    break
                                else:
                                    found_data =False
                                    driver.back()
                            elif tab == 'Archived':
                                archived_client_id_element = WebDriverWait(driver,120).until(EC.visibility_of_element_located((By.XPATH,f"//table//tbody//tr[{tbl_row}]/td[3]"))) #//table//tbody//tr[1]/td[3]
                                archived_client_id_number = archived_client_id_element.text
                                if client_id_number == archived_client_id_number:
                                    archived_tab = True
                                    unarchive_button_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,f"(//button[@data-aqa='btnUnarchive'])[{tbl_row}]")))
                                    unarchive_button_element.click()

                                    submit_button_element = WebDriverWait(driver,120).until(EC.element_to_be_clickable((By.XPATH,'//button[@data-aqa="btnSubmit"]')))
                                    submit_button_element.click()
                                    time.sleep(1)

                                    for attempt in range(3):
                                        try:
                                            tab_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), 'Active')])[1]")))
                                            time.sleep(1)
                                            tab_element.click()
                                            time.sleep(2)
                                            break
                                        except StaleElementReferenceException:
                                            continue

                                    for name in name_to_search:
                                        search_bar_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
                                        search_bar_element.clear()
                                        search_bar_element.send_keys(name)
                                        time.sleep(2)
                                        try:
                                            # driver.find_element(By.XPATH, "//div[text()='No data available']")
                                            no_data_element = WebDriverWait(driver,120).until(EC.visibility_of_element_located((By.XPATH,"//div[text()='No data available']")))
                                            search_bar_element.clear()
                                        except:
                                            found_data = True
                                            clientname_table_element = WebDriverWait(driver, 120).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[contains(@class, 'k-grid-table')]//tr")))
                                            for tbl_row in range(1, len(clientname_table_element) + 1):
                                                time.sleep(2)
                                                for attempt in range(3):
                                                    try:
                                                        client_name_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                                        time.sleep(1)
                                                        client_name_element.click()
                                                        break
                                                    except StaleElementReferenceException:
                                                        continue
                                                client_id_element = WebDriverWait(driver, 40).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-aqa='ClientID']//div[2]")))
                                                client_id_element_number = client_id_element.text

                                                if client_id_number != client_id_element_number:
                                                    found_data = False
                                                    archived_tab = False
                                                    driver.back()
                                                else:
                                                    claim_url = driver.current_url
                                                    firstname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='FirstName']"))).get_attribute("value")
                                                    middlename = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='MiddleName']"))).get_attribute("value")
                                                    lastname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='LastName']"))).get_attribute("value")

                                                    portal_client_name = f"{firstname} {middlename} {lastname}".strip()
                                                    portal_client_name = " ".join(portal_client_name.split())
                                                    break
                                        if archived_tab:
                                            break
                            if found_data:
                                break
                        if archived_tab or found_data:
                            break
                if found_data:
                    break

            if not found_data:
                print(f"❌ Patient: {client_name} not found.!")
                scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value = "Client Name or ID not found"
                sf_billing_wb.save(billing_file_path)
                continue

            if active_tab or claimTab_status != "Archived":
                scrubbing_sheet.cell(row=row, column=data_columns['Active/Archived']).value = "Active"
                sf_billing_wb.save(billing_file_path)

            elif archived_tab:
                scrubbing_sheet.cell(row=row, column=data_columns['Active/Archived']).value = "Archived"
                sf_billing_wb.save(billing_file_path)


            ### Navigate to ledger ###
            ledger_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"//a[@aria-label='Ledger']")))
            ledger_btn_element.click()

            newinvoice_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"//button[@data-aqa='btnNewInvoice']")))
            newinvoice_btn_element.click()

            recent_btn = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnRecentDays']")))
            recent_btn.click()

            # all_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"//button[@data-aqa='btnAll']")))
            # all_btn_element.click()
            time.sleep(2)

            ### Services Table looping ###
            dos_match = False
            extracted_dos_date = dos_date.replace(" ET", "").strip()
            service_table_element = WebDriverWait(driver, 120).until(EC.presence_of_all_elements_located((By.XPATH,"(//table[contains(@class, 'k-grid-table')])[1]//tr")))
            driver.execute_script("document.body.style.zoom= '50%'")
            if len(service_table_element) >= 1 :
                for tbl_row in range(1, len(service_table_element) +1):

                    if dos_match:
                        tbl_row = 2
                    else:
                        tbl_row = 1

                    service_tblrow_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,f"(//table[contains(@class, 'k-grid-table')]//tr)[{tbl_row}]//a")))
                    service_dos_txt = service_tblrow_element.text
                    if status == "Late Cancel" or status == "No Show":
                        service_type_element = driver.find_element(By.XPATH, f"(//table[contains(@class, 'k-grid-table')]//tr[{tbl_row}]//select)[1]")
                        title_value = service_type_element.get_attribute("title")
                        if extracted_dos_date == service_dos_txt and service_type == title_value:
                            dos_match = True
                            print(f"DOS: ({extracted_dos_date}) matched.")
                        else:
                            claim_remove_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"(//button[@data-aqa='removeService'])[{tbl_row}]")))
                            actions.move_to_element(claim_remove_element).perform()
                            time.sleep(1)
                            claim_remove_element.click()
                            time.sleep(1)

                    else:
                        if extracted_dos_date == service_dos_txt:
                            dos_match = True
                            print(f"DOS: ({extracted_dos_date}) matched.")
                        else:
                            claim_remove_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"(//button[@data-aqa='removeService'])[{tbl_row}]")))
                            actions.move_to_element(claim_remove_element).perform()
                            # driver.execute_script("arguments[0].scrollIntoView(true);", claim_remove_element)
                            time.sleep(1)
                            claim_remove_element.click()
                            time.sleep(1)

            else:
                print(f"❌ DOS: ({extracted_dos_date}) does not match. No claim entry found in the service table.")
                scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value = "No claims are present in particular DOS"
                sf_billing_wb.save(billing_file_path)
                continue

            provider_warning = False
            move_awaiting = False
            if dos_match:

                dxcode_check_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"(//button[@data-aqa='editServicePqrsCodes'])[1]")))
                dxcode_check_element.click()
                time.sleep(1)

                daigonis_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,"//form//div[@class='content']//p")))
                daignosis_text = daigonis_element.text.strip()
                print(daignosis_text)
                # time.sleep(2)
                close_dxTab_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='ModalCloseButton']")))
                close_dxTab_element.click()
                # no_dx_reason = ["None of the Cases related to this session have billable diagnoses entered.", "No diagnoses found because the session is not associated with a progress note."]
                # if any(dx_reason in daignosis_text for dx_reason in no_dx_reason):
                no_dx_label = "None of the Cases related to this session have billable diagnoses entered."
                if not no_dx_label in daignosis_text:
                    # Set Copay amount $0.00
                    if status == "Late Cancel" or status == "No Show":
                        amount_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceAmount'])[1]")))
                        amount_element.clear()
                        amount_element.send_keys("65.00")

                        copay_amount_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceClientAmountDue'])[1]")))
                        copay_amount_element.clear()
                        copay_amount_element.send_keys("65.00")

                        insurance_amount_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceInsuranceAmountDue'])[1]")))
                        insurance_amount_element.clear()
                        insurance_amount_element.send_keys("0.00")
                        time.sleep(1)
                    else:

                        copay_amount_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceClientAmountDue'])[1]")))
                        copay_amount_element.clear()
                        copay_amount_element.send_keys("0.00")

                        units_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@data-aqa='serviceUnits']")))
                        units_element.clear()
                        units_element.send_keys("1")
                        time.sleep(1)


                    if availity_payor:
                        ### capture the DX Code for availity Portal Purpose
                        dxcodes = re.findall(r"\(([^)]+)\)", daignosis_text)
                        merged_dxcodes = ",".join(dxcodes)

                        ### Claim_DOS details Scrubbing ###
                        placeofservice_element = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.XPATH, "(//select[@data-aqa='servicePlaceOfService'])[1]")))
                        place_of_service = placeofservice_element.get_attribute("value")
                        # place_of_service = re.findall(r"\(([^)+])\)", place_of_service)

                        patient_amount_element = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceAmount'])[1]")))
                        insurance_amount = patient_amount_element.get_attribute("value")

                        patient_unit_element = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.XPATH, "(//input[@data-aqa='serviceUnits'])[1]")))
                        unit = patient_unit_element.get_attribute("value")



                        if payor_name in availitypayor_staffmember_dict:
                            if staff_member in availitypayor_staffmember_dict:
                                rendering_provider = availitypayor_staffmember_dict[staff_member]['Rendering Provider']
                                billing_provider = availitypayor_staffmember_dict[staff_member]['Billing Provider']
                            else:
                                print(f"The Staff name ({staff_member}) is missing in the Staff Member table masters. Kindly check and re run the script")
                                sys.exit()
                    else:
                        if payor_name in staff_member_payors:
                            if staff_member in staffmemberwise_providerDf_dict:
                                rendering_provider = staffmemberwise_providerDf_dict[staff_member]['Rendering Provider']
                                billing_provider = staffmemberwise_providerDf_dict[staff_member]['Billing Provider']
                            else:
                                print(f"The Payor name ({staff_member}) is missing in the Staff Member table masters. Kindly check and re run the script")
                                sys.exit()
                        else:
                            if payor_name in payorwise_providerDf_dict:
                                rendering_provider = payorwise_providerDf_dict[payor_name]['Rendering Provider']
                                billing_provider = payorwise_providerDf_dict[payor_name]['Billing Provider']
                            else:
                                print(f"The Payor name ({payor_name}) is missing in the Payor table masters. Kindly check and re run the script")
                                sys.exit()

                    # Set Rendering & Billing Provider
                    renderingProvider_dropdown = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//div[@data-aqa='invoiceRenderingProvider']")))
                    renderingProvider_dropdown.click()

                    rendering_provider_option = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH,f"(//div[@data-aqa='invoiceRenderingProvider']//div[@role='option' and contains(normalize-space(), '{rendering_provider}')])")))
                    time.sleep(1)
                    actions.move_to_element(rendering_provider_option).click().perform()


                    if billing_provider != "No":
                        bill_checkbox = WebDriverWait(driver, 120).until(
                            EC.presence_of_element_located((By.XPATH,"//span[@data-aqa='isBillWithStaffNpi']")))
                        try:
                            bill_checkbox.find_element(By.TAG_NAME, "path")
                            is_checked = True
                        except NoSuchElementException:
                            is_checked = False

                        if is_checked:
                            billingProvider_dropdown = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//div[@data-aqa='invoiceBillingProvider']")))
                            billingProvider_dropdown.click()

                            billing_provider_option = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,f"(//div[@data-aqa='invoiceBillingProvider']//div[@role='option' and contains(normalize-space(), '{billing_provider}')])")))
                            time.sleep(1)
                            actions.move_to_element(billing_provider_option).click().perform()
                        else:
                            bill_checkbox.click()
                            time.sleep(1)

                            billingProvider_dropdown = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//div[@data-aqa='invoiceBillingProvider']")))
                            billingProvider_dropdown.click()

                            billing_provider_option = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,f"(//div[@data-aqa='invoiceBillingProvider']//div[@role='option' and contains(normalize-space(), '{billing_provider}')])")))
                            time.sleep(1)
                            actions.move_to_element(billing_provider_option).click().perform()

                    # Move to awaiting
                    if status not in ["Late Cancel", "No Show"] and not availity_payor:
                        save_dropdown_btn_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH,"(//div[@data-aqa='btnSaveOptions'])[1]")))
                        time.sleep(1)
                        actions.move_to_element(save_dropdown_btn_element).perform()
                        time.sleep(1)
                        save_dropdown_btn_element.click()
                        time.sleep(1)

                        claimAwaiting_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "(//div[@data-aqa='addToClaimsAwaitingSubmission'])[1]")))
                        claimAwaiting_element.click()
                        time.sleep(1)

                        move_awaiting = True

                else:
                    scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value = "DX code not found"
                    sf_billing_wb.save(billing_file_path)
                    continue

            else:
                print(f"❌ DOS: ({extracted_dos_date}) does not match. No claim entry found in the service table.")
                scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value = "No claims are present in particular DOS"
                continue
            if status == "Late Cancel" or status == "No Show":
                save_element = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "(//button[@data-aqa='saveInvoice'])[1]")))
                actions.move_to_element(save_element).click().perform()
                time.sleep(1)
                scrubbing_sheet.cell(row=row, column=data_columns['Is Billed']).value = "Yes"
                sf_billing_wb.save(billing_file_path)
                print(f"✅ Billing for patient ({client_name}) has been successfully moved to insurance.")

                if availity_payor:
                    client_url_number = claim_url.split('/')[-1]

                    ### Invoice Number ###
                    openInvoice_tbl_element = WebDriverWait(driver, 120).until(
                        EC.presence_of_all_elements_located((By.XPATH, "(//table[@class='k-grid-table'])[1]")))

                    invoicenumber_element = WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, "(//table//td[@data-aqa='number'])[1]")))
                    invoice_number = invoicenumber_element.text

                    ### ClientDetails Scrubbing ###
                    driver.get(f"https://wisemind71.theranest.com/clients/details/{client_url_number}")

                    patient_dob_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@name='DemographicInfoBirthDate']")))
                    dob = patient_dob_element.get_attribute("value")

                    patient_gender_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//select[@name='DemographicInfoGender']")))
                    gender = patient_gender_element.get_attribute("value")

                    patient_street_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@name='AddressStreet']")))
                    street = patient_street_element.get_attribute("value")

                    patient_city_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@name='AddressCity']")))
                    city = patient_city_element.get_attribute("value")

                    patient_state_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//select[@name='AddressState']")))
                    state = patient_state_element.get_attribute("value")

                    patient_zipcode_element = WebDriverWait(driver, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@name='AddressZip']")))
                    zip_code = patient_zipcode_element.get_attribute("value")

                    ### Insurance details Scrubbing ###
                    driver.get(f"https://wisemind71.theranest.com/clients/billing-info/{client_url_number}")

                    insurance_table_element = WebDriverWait(driver, 60).until(
                        EC.presence_of_all_elements_located((By.XPATH, "(//table[@class='k-grid-table'])[1]//tr")))

                    if len(insurance_table_element):
                        for tbl_row in range(1, len(insurance_table_element)+1):

                            payorname_row_element = WebDriverWait(driver, 60).until(
                                EC.presence_of_element_located((By.XPATH, f"(//table[@class='k-grid-table'])[1]//tr{tbl_row}//td[@data-aqa='provider']")))
                            payorname = payorname_row_element.get_attribute("value")

                            payor_status_element = WebDriverWait(driver, 60).until(
                                EC.presence_of_element_located((By.XPATH, f"(//table[@class='k-grid-table'])[1]//tr{tbl_row}//td[@data-aqa='status']//div[@class='azk_cy azk_a1 azk_kl']")))
                            payor_status = payor_status_element.get_attribute("value")

                            if payorname in availitypayor and payor_status == "Active":
                                insurance_id_check = True
                                insurance_id_element = WebDriverWait(driver, 60).until(
                                    EC.visibility_of_element_located((By.XPATH, f"(//table[@class='k-grid-table'])[1]//tr{tbl_row}//td[@data-aqa='insuredIdNumber']")))
                                insurance_id = insurance_id_element.get_attribute("value")
                                break
                            else:
                                continue

                    if availity_payor and insurance_id_check:
                        if not os.path.exists(bcbs_file_path):
                            bcbs_logbook = openpyxl.workbook()
                            bcbs_logsheet = bcbs_logbook.active

                            headers = ["Client Name", "Client ID Number", "Date/Time", "Service Type", "Staff Member(s)", "Payor Name", "First Name", "Last Name", "DOB", "Gender", "Street", "City", "State", "ZIP Code", "Claim Invoice Number", "Place of Service", "DX Code", "Charge Amount","Quantity"]
                            for col_num, headers in enumerate(headers, start=1):
                                bcbs_logsheet.cells(row=1, column=col_num, value=headers)

                            bcbs_logbook.save(bcbs_file_path)
                        else:
                            bcbs_logbook = load_workbook(bcbs_file_path)
                            bcbs_logsheet = bcbs_logbook.active

                            bcbs_data_column = {}
                            for col in range(1, bcbs_logsheet.max_column + 1):
                                col_name = bcbs_logsheet.cell(row=1, column=col).value
                                if col_name:
                                    bcbs_data_column[col_name.strip()] = col

                            max_row = bcbs_logsheet.max_row + 1

                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Client Name"]).value = client_name
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Client ID Number"]).value = client_id_number
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Date/Time"]).value = dos_date
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Service Type"]).value = service_type
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Staff Member(s)"]).value = staff_member
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Payor Name"]).value =payor_name
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["First Name"]).value = firstname
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Last Name"]).value = lastname
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["DOB"]).value = dob
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Gender"]).value = gender
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Street"]).value = street
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["City"]).value = city
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["State"]).value = state
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["ZIP Code"]).value = zip_code
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Claim Invoice Number"]).value = invoice_number
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Place of Service"]).value = place_of_service
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["DX Code"]).value = dxcodes
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Charge Amount"]).value = insurance_amount
                            bcbs_logsheet.cell(row=max_row, column=bcbs_data_column["Quantity"]).value = unit

                            bcbs_logbook.save(bcbs_file_path)

                    else:
                        driver.get(f"https://wisemind71.theranest.com/ledger/client-{client_url_number}/open-invoices")
                        time.sleep(10)

                        openInvoice_tbl_element = WebDriverWait(driver, 120).until(
                            EC.presence_of_all_elements_located((By.XPATH, "(//table[@class='k-grid-table'])[1]")))

                        action_btn_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, "//table//tbody//tr[1]//td[7]//div[@data-aqa='btnActions']")))
                        action_btn_element.click()

                        delete_invoiceBtn_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, "//table//tbody//tr[1]//td[7]//div[@data-aqa='Delete']")))
                        delete_invoiceBtn_element.click()

                        delete_invoiceBtn2_element = WebDriverWait(driver, 60).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnDeleteInvoice']")))
                        delete_invoiceBtn2_element.click()
                        delete_invoiceBtn2_element = WebDriverWait(driver, 60).until(
                            EC.invisibility_of_element_located((By.XPATH, "//button[@data-aqa='btnDeleteInvoice']")))
                        time.sleep(1)

                        scrubbing_sheet.cell(row=row, column=data_columns[
                            'Exceptions']).value = "The claim encountered an error during the 'Awaiting Submission' process."
                        sf_billing_wb.save(billing_file_path)
                        print(f"❌ Patient ({client_name}) has no active BCBS insurance or missing Insurance ID, so the claim was moved to exceptions.")
                        print("\n")




            if move_awaiting:
                try:
                    awaiting_searchbar_element = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "(//div[@data-aqa='inputClientFullName']//input)")))
                except:
                    try:
                        choose_insurance_element = WebDriverWait(driver,5).until(EC.presence_of_element_located((By.XPATH, "//button[@data-aqa='usePrimarySubcsription']")))
                        choose_insurance_element.click()
                    except:
                        pass
                    try:
                        rendering_provider_warning_element = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='ModalOkButton']")))
                        rendering_provider_warning_element.click()
                        provider_warning = True
                        time.sleep(1)
                    except:
                        pass

                awaiting_searchbar_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "(//div[@data-aqa='inputClientFullName']//input)")))
                awaiting_searchbar_element.clear()
                awaiting_searchbar_element.send_keys(portal_client_name)
                time.sleep(2)

                try:
                    no_data_element = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='yg_yk']")))
                    awaiting_searchbar_element.clear()
                except:
                    awaiting_tbl_element = WebDriverWait(driver, 120).until(EC.presence_of_all_elements_located((By.XPATH, "//table[@class='k-grid-table']//tr")))
                    if len(awaiting_tbl_element) > 1:
                        for tbl_row in range(1, len(awaiting_tbl_element) +1):
                            for attempt in range(3):
                                try:
                                    client_name_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                    time.sleep(1)
                                    client_name_element.click()
                                    break
                                except StaleElementReferenceException:
                                    continue
                            client_id_element = WebDriverWait(driver, 40).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-aqa='ClientID']//div[2]")))
                            client_id_element_number = client_id_element.text

                            if client_id_number != client_id_element_number:
                                driver.back()
                            else:
                                driver.back()
                                awaiting_client_checkbox_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[7]")))
                                awaiting_client_checkbox_element.click()
                                actions.send_keys(Keys.TAB).perform()
                                time.sleep(1)
                                actions.send_keys(Keys.SPACE).perform()
                                # if not awaiting_client_checkbox_element.is_enabled():
                                #     awaiting_client_checkbox_element.click()

                                error_check_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnCheckSelectedClaimsforErrors']")))
                                error_check_btn_element.click()

                                if provider_warning:
                                    try:
                                        confirm_action_btn_element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnContinue']")))
                                        confirm_action_btn_element.click()
                                    except:
                                        pass
                                break
                    else:
                        awaiting_client_checkbox_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//table//tbody//tr[1]/td[7]")))
                        awaiting_client_checkbox_element.click()
                        time.sleep(1)
                        actions.send_keys(Keys.TAB).perform()
                        time.sleep(1)
                        actions.send_keys(Keys.SPACE).perform()
                        # actions.move_to_element(awaiting_client_checkbox_element).click().perform()
                        # awaiting_client_checkbox_element.click()
                        # checkbox_label = WebDriverWait(driver, 20).until(
                        #     EC.element_to_be_clickable((By.CSS_SELECTOR, "label.k-checkbox-label.amq_amr"))
                        # )
                        # checkbox_label.click()
                        time.sleep(1)
                        error_check_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnCheckSelectedClaimsforErrors']")))
                        error_check_btn_element.click()

                        if provider_warning:
                            try:
                                confirm_action_btn_element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnContinue']")))
                                confirm_action_btn_element.click()
                            except:
                                pass

                        try:
                            final_error_check_element = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//a[@id='close_errors']")))
                            final_error_check_element.click()

                            remove_claim_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnRemoveSelectedClaims']")))
                            remove_claim_element.click()
                            time.sleep(1)

                            client_url_number = claim_url.split('/')[-1]
                            driver.get(f"https://wisemind71.theranest.com/ledger/client-{client_url_number}/open-invoices")
                            time.sleep(10)

                            openInvoice_tbl_element = WebDriverWait(driver, 120).until(
                                EC.presence_of_all_elements_located((By.XPATH, "(//table[@class='k-grid-table'])[1]")))

                            action_btn_element = WebDriverWait(driver, 60).until(
                                EC.element_to_be_clickable((By.XPATH, "//table//tbody//tr[1]//td[7]//div[@data-aqa='btnActions']")))
                            action_btn_element.click()

                            delete_invoiceBtn_element = WebDriverWait(driver, 60).until(
                                EC.element_to_be_clickable((By.XPATH, "//table//tbody//tr[1]//td[7]//div[@data-aqa='Delete']")))
                            delete_invoiceBtn_element.click()

                            delete_invoiceBtn2_element = WebDriverWait(driver, 60).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnDeleteInvoice']")))
                            delete_invoiceBtn2_element.click()
                            delete_invoiceBtn2_element = WebDriverWait(driver, 60).until(
                                EC.invisibility_of_element_located((By.XPATH, "//button[@data-aqa='btnDeleteInvoice']")))
                            time.sleep(1)

                            scrubbing_sheet.cell(row=row, column=data_columns['Exceptions']).value = "The claim encountered an error during the 'Awaiting Submission' process."
                            sf_billing_wb.save(billing_file_path)
                            print(f"❌ patient ({client_name}) have Awaiting error.")
                            print("\n")

                        except:
                            # if status == "Late Cancel":
                            #     save_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "(//button[@data-aqa='saveInvoice'])[1]")))
                            #     actions.move_to_element(save_element).click().perform()
                            #     time.sleep(1)
                            #     scrubbing_sheet.cell(row=row, column=data_columns['Is Billed']).value = "Yes"
                            #     sf_billing_wb.save(billing_file_path)
                            #     print(f"✅ Billing for patient ({client_name}) has been successfully moved to insurance.")
                            # else:

                            submit_claim_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnSubmitSelectedClaims']")))
                            actions.move_to_element(submit_claim_element).click().perform()
                            scrubbing_sheet.cell(row=row, column=data_columns['Is Billed']).value = "Yes"
                            sf_billing_wb.save(billing_file_path)
                            print(f"✅ Billing for patient ({client_name}) has been successfully moved to insurance.")
                            print("\n")


            ### If the claim has been moved from 'Archived' to 'Active', we now need to move it back to 'Archived' ###
            if archived_tab or claimTab_status == "Archived":

                driver.get("https://wisemind71.theranest.com/clients")
                for attempt in range(3):
                    try:
                        tab_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), 'Active')])[1]")))
                        tab_element.click()
                        time.sleep(2)
                        break
                    except StaleElementReferenceException:
                        time.sleep(1)
                        continue
                search_bar_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
                search_bar_element.clear()
                search_bar_element.send_keys(portal_client_name)
                time.sleep(1)
                try:
                    no_data_element = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//div[text()='No data available']")))
                    search_bar_element.clear()
                except:
                    clientname_table_element = WebDriverWait(driver, 120).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[contains(@class, 'k-grid-table')]//tr")))
                    if len(clientname_table_element) > 1:
                        for tbl_row in range(1, len(clientname_table_element) + 1):
                            time.sleep(2)
                            for attempt in range(3):
                                try:
                                    client_name_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                    time.sleep(1)
                                    client_name_element.click()
                                    break
                                except StaleElementReferenceException:
                                    continue
                            client_id_element = WebDriverWait(driver, 40).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-aqa='ClientID']//div[2]")))
                            client_id_element_number = client_id_element.text

                            if client_id_number != client_id_element_number:
                                driver.back()
                            else:
                                driver.back()
                                archived_btn_element = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, f"(//button[@data-aqa='btnArchive'])[{tbl_row}]")))
                                archived_btn_element.click()

                                arcSubmit_element = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnSubmit']")))
                                arcSubmit_element.click()
                                break
                    else:
                        archived_btn_element = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "(//button[@data-aqa='btnArchive'])[1]")))
                        archived_btn_element.click()

                        arcSubmit_element = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnSubmit']")))
                        arcSubmit_element.click()

        time.sleep(2)
        driver.get("https://wisemind71.theranest.com/home/logout")
        time.sleep(3)

        print("✅✅ Billing Completed! ✅✅")
        driver.quit()


