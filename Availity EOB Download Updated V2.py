import getpass
import glob
import os
import re
import shutil
import sys
import pandas as pd
import openpyxl
import numpy as np
import time
from openpyxl.workbook import Workbook
from pywinauto.keyboard import send_keys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta


todayDate = datetime.today().strftime('%m%d%Y')
todayDate1 = datetime.today().strftime("%m/%d/%Y")
current_month = datetime.now().strftime("%B")[0:3]


start_of_month = datetime.today().replace(day=1)
start_of_month_formatted = start_of_month.strftime("%m/%d/%Y")

yesterday = datetime.now() - timedelta(days=1)
yesterday_formatted = yesterday.strftime("%m/%d/%Y")

systemUsername = getpass.getuser()
downloads_path = fr"C:\Users\{systemUsername}\Downloads"
list_of_files = glob.glob(f"{downloads_path}/*")
start_latest_file = max(list_of_files, key=os.path.getmtime)


configWB = openpyxl.load_workbook(r"Z:\Rethink\Payment Posting\09. EOB\Automation Config Files\Availity Master Config Sheet.xlsx")
configSheet = configWB['Credentials']
username = configSheet.cell(row=1, column=2).value
password = configSheet.cell(row=2, column=2).value
websiteURL = configSheet.cell(row=3, column=2).value
remittanceURL = configSheet.cell(row=4, column=2).value
sharedParentPath = configSheet.cell(row=5, column=2).value
homePageURL = configSheet.cell(row=6, column=2).value


if not os.path.exists(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]):
    os.makedirs(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:])
if not os.path.exists(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate):
    os.makedirs(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate)
if not os.path.exists(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate+"\\"+"Availity EOB"):
    os.makedirs(sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate+"\\"+"Availity EOB")

filename = sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate+"\\"+"Availity EOB\\"+f"Availity EOB-{todayDate}.xlsx"

if not os.path.exists(filename):
    configDf = pd.read_excel(r"Z:\Rethink\Payment Posting\09. EOB\Automation Config Files\Availity Master Config Sheet.xlsx",sheet_name='Details')
    configDf = configDf.replace(np.nan, '')
    configDf['Status'] = ''
    configDf.to_excel(filename, index=False)
    inputDf = pd.read_excel(filename)
    inputDf = inputDf.replace(np.nan, '')
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['Sheet1']
else:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['Sheet1']
    inputDf = pd.read_excel(filename)
    inputDf = inputDf.replace(np.nan, '')


chequeDetailsFilename = sharedParentPath+todayDate[0:2]+" "+current_month+"'"+todayDate[4:]+"\\"+todayDate+"\\"+"Availity EOB\\"+f"\\ChequeNumberDetails-{todayDate}.xlsx"
if not os.path.exists(chequeDetailsFilename):
    chequeWorkbook = Workbook()
    chequeWorksheet = chequeWorkbook.active
    chequeWorksheet.title = 'Sheet1'
    header = ["Check/EFT", "Payer", "Payee", "Check Date", "Check/EFT Amount", "Status", "RSM"]
    for col_num, header_title in enumerate(header, start=1):
        chequeWorksheet.cell(row=1, column=col_num, value=header_title)
    chequeWorkbook.save(chequeDetailsFilename)
else:
    chequeWorkbook = openpyxl.load_workbook(chequeDetailsFilename)
    chequeWorksheet = workbook['Sheet1']
df = pd.read_excel(chequeDetailsFilename)
df = df.replace(np.nan,'')
length = len(df)
uniqueCheckList = df['Check/EFT'].unique().tolist()




chrome_options = Options()
chrome_options.add_argument("--start-maximized")
chrome_options.add_experimental_option('detach', True)


driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.get(websiteURL)
driver.maximize_window()
action = ActionChains(driver)
time.sleep(5)
usernameElement = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.NAME, "userId")))
usernameElement.send_keys(username)
time.sleep(2)
passwordElement = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.NAME, "password")))
passwordElement.send_keys(password)
time.sleep(2)
signInElement = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, "//button[normalize-space()='Sign In']")))
signInElement.click()
time.sleep(2)



code = input("Please enter the code manually and hit enter here:")
time.sleep(7)
popupCheck = input("Please clear all pop content then hit enter if not pop up content then go ahead and hit enter:")
time.sleep(7)
if code.strip() == '' and popupCheck == '':
    for index, row in inputDf.iterrows():
        if username == row['Availity Login']:
            driver.get(homePageURL)
            time.sleep(8)
            stateName = row['State Name'].strip()
            rsmName = row['RSMs']
            practiceName = row['Practice Name']
            if row['Status'] == '' :
                stateElement = WebDriverWait(driver, 120).until(
                    EC.visibility_of_element_located((By.XPATH, "//span[@id='select2-chosen-1']")))
                stateNameFromPortal = stateElement.text
                # print(stateNameFromPortal)
                # print(stateName)
                if stateNameFromPortal.strip() != stateName:
                    stateElement.click()
                    time.sleep(2)
                    stateSearchBoxElement = WebDriverWait(driver, 120).until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@id='s2id_autogen1_search']")))
                    stateSearchBoxElement.send_keys(stateName)
                    time.sleep(2)
                    searchBoxResultElement = WebDriverWait(driver, 120).until(
                        EC.visibility_of_element_located((By.XPATH, "(//ul[@id='select2-results-1'])[1]")))
                    stateNameFromPortal = searchBoxResultElement.text
                    searchBoxResultElement.click()
                    time.sleep(5)
                if stateName == stateNameFromPortal.strip():
                    stateElementCount = 0
                    while stateElementCount < 60:
                        stateElement = WebDriverWait(driver, 120).until(
                            EC.visibility_of_element_located((By.XPATH, "//span[@id='select2-chosen-1']")))
                        if stateElement.text == stateName:
                            break
                        else:
                            time.sleep(1)
                            stateElementCount+=1
                    if stateElementCount >= 60:
                        print("Selecting state is for so long time please check")
                        sys.exit()
                    driver.get(remittanceURL)
                    time.sleep(10)
                    if popupCheck.strip() == "":
                        frameElement = WebDriverWait(driver, 60).until(
                            EC.frame_to_be_available_and_switch_to_it((By.ID, "newBodyFrame")))
                        time.sleep(2)
                        practiceElement = WebDriverWait(driver, 120).until(
                            EC.visibility_of_element_located((By.XPATH, "//*[@id='organizationId']")))
                        action.move_to_element(practiceElement).click().perform()
                        time.sleep(2)
                        practiceElement.send_keys(practiceName)
                        time.sleep(3)
                        send_keys('{ENTER}')

                        # receiverStartDateElement = WebDriverWait(driver, 120).until(
                        #     EC.visibility_of_element_located((By.XPATH, "//*[@id='checkEFTcheckExchangeDates-start']")))
                        # time.sleep(1)
                        # action.move_to_element(receiverStartDateElement).click().click().click().perform()
                        # time.sleep(1)
                        # receiverStartDateElement.send_keys(yesterday_formatted)
                        #
                        # receiverEndDateElement = WebDriverWait(driver, 120).until(
                        #     EC.visibility_of_element_located((By.XPATH, "//*[@id='checkEFTcheckExchangeDates-end']")))
                        # time.sleep(1)
                        # action.move_to_element(receiverEndDateElement).click().click().click().perform()
                        # time.sleep(1)
                        # receiverEndDateElement.send_keys(todayDate1)

                        filterButtonElement = WebDriverWait(driver, 120).until(
                            EC.visibility_of_element_located((By.XPATH, "//*[@id='checkFilterButton']")))
                        filterButtonElement.click()
                        time.sleep(3)
                        checkPracticeElement = WebDriverWait(driver, 120).until(
                            EC.visibility_of_element_located((By.XPATH, '//*[@id="checkEFTorganizationId"]/div/div[1]/div[1]')))
                        print(checkPracticeElement.text)
                        if checkPracticeElement.text == 'All':
                            worksheet.cell(row=index + 2, column=6).value = "Practice name not found in Drop Down"
                            workbook.save(filename)
                            continue
                        startDateElement = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.ID, "checkcheckDates-start")))
                        time.sleep(1)
                        action.move_to_element(startDateElement).click().click().click().perform()
                        time.sleep(1)
                        startDateElement.send_keys(yesterday_formatted)
                        time.sleep(1)

                        endDateElement = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.ID, "checkcheckDates-end")))
                        time.sleep(1)
                        action.move_to_element(endDateElement).click().click().click().perform()
                        time.sleep(1)
                        endDateElement.send_keys(todayDate1)
                        time.sleep(1)

                        searchElement = WebDriverWait(driver, 60).until(
                            EC.visibility_of_element_located((By.ID, "checkSearchButton")))
                        searchElement.click()
                        time.sleep(1)
                        try:
                            paginatorElement = WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                                (By.XPATH, "//div[@class='align-items-center row']//span[@class='mx-2 mt-2']")))
                            paginatorText = paginatorElement.text
                            paginatorText = paginatorText.strip()
                        except:
                            worksheet.cell(row=index + 2, column=6).value = "No EOBs for Cheque Date"
                            workbook.save(filename)
                            continue
                        # print(paginatorText)
                        startCount = int(paginatorText.split(' ')[1])
                        endCount = int(paginatorText.split(' ')[3])
                        countPerPage = (endCount - startCount) + 1
                        if 'more' not in paginatorText:
                            totalCount = int(paginatorText.split(' ')[5])
                        elif 'more' in paginatorText:
                            totalCount = int(paginatorText.split(' ')[7])

                        if endCount == totalCount:
                            for i in range(countPerPage):
                                rowElement = WebDriverWait(driver, 60).until(
                                    EC.visibility_of_all_elements_located((By.XPATH, f"//*[@id='checkNumber{i}']")))
                                rowText = rowElement[0].text

                                checkNumber = rowText.split('\n')[0]
                                payer = re.sub(r"[:?/\|''*<>]", "", rowText.split('\n')[1])
                                payee = re.sub(r"[:?/\|''*<>]", "", rowText.split('\n')[2])
                                checkDate = rowText.split('\n')[3]
                                receivedByAvailityDate = rowText.split('\n')[4]
                                checkAmt = rowText.split('\n')[5]

                                # print(checkNumber)
                                # print(payer)
                                # print(payee)
                                # print(checkAmt)
                                if receivedByAvailityDate == todayDate1 or receivedByAvailityDate == yesterday_formatted or checkDate == todayDate1 or checkDate == yesterday_formatted:
                                    try:
                                        if checkNumber not in uniqueCheckList:
                                            pdfIconElement = WebDriverWait(driver, 5).until(
                                                EC.visibility_of_element_located((By.ID, f"checkeob{i}")))
                                            action.move_to_element(pdfIconElement).click().perform()
                                            # pdfIconElement.click()
                                        else:
                                            continue
                                        time.sleep(2)
                                        try:
                                            checkboxElement = WebDriverWait(driver, 10).until(
                                                EC.visibility_of_element_located(
                                                    (By.XPATH, f"//label[@for='allcheckeob{i}']")))
                                            checkboxElement.click()
                                            time.sleep(2)
                                            downloadElement = WebDriverWait(driver, 10).until(
                                                EC.element_to_be_clickable(
                                                    (By.XPATH, "//button[normalize-space()='Download']")))
                                            downloadElement.click()
                                        except:
                                            pass
                                        time.sleep(8)
                                        downloadWaitCount = 0
                                        while downloadWaitCount < 120:
                                            list_of_files = glob.glob(f"{downloads_path}/*")
                                            try:
                                                latest_file = max(list_of_files, key=os.path.getmtime)
                                            except FileNotFoundError:
                                                time.sleep(1)
                                                downloadWaitCount += 1
                                                continue
                                            if latest_file != start_latest_file and latest_file.endswith('.pdf'):
                                                shutil.move(latest_file, sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf")
                                                break
                                            else:
                                                time.sleep(1)
                                                downloadWaitCount += 1
                                                continue
                                        if downloadWaitCount >= 120:
                                            print(
                                                f'The File with this check number --{checkNumber}--  is being downloaded for so long time please check')
                                            try:
                                                closeElement = WebDriverWait(driver, 10).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, "//button[normalize-space()='Close']")))
                                                closeElement.click()
                                            except:
                                                pass
                                            try:
                                                cancelElement = WebDriverWait(driver, 10).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, "//*[@class='icon icon-cancel cursor-pointer']")))
                                                cancelElement.click()
                                            except:
                                                pass
                                            # sys.exit()
                                        if os.path.exists(sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf"):
                                            chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                            chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                            chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                            chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                            chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                            chequeWorksheet.cell(row=length + 2, column=6).value = "File Downloaded"
                                            try:
                                                chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                            except KeyError:
                                                chequeWorksheet.cell(row=length + 2,
                                                               column=7).value = 'Payee not found in Master'
                                            chequeWorkbook.save(chequeDetailsFilename)
                                            df = pd.read_excel(chequeDetailsFilename)
                                            df = df.replace(np.nan, '')
                                            length = len(df)
                                        else:
                                            chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                            chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                            chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                            chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                            chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                            chequeWorksheet.cell(row=length + 2, column=6).value = "File Not Downloaded"
                                            try:
                                                chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                            except KeyError:
                                                chequeWorksheet.cell(row=length + 2,
                                                               column=7).value = 'Payee not found in Master'
                                            chequeWorkbook.save(chequeDetailsFilename)
                                            df = pd.read_excel(chequeDetailsFilename)
                                            df = df.replace(np.nan, '')
                                            length = len(df)
                                        try:
                                            closeElement = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                                                (By.XPATH, "//button[normalize-space()='Close']")))
                                            closeElement.click()
                                        except:
                                            pass
                                    except:
                                        if checkNumber not in uniqueCheckList:
                                            actionsElement = WebDriverWait(driver, 60).until(
                                                EC.visibility_of_element_located((By.XPATH,
                                                                                  f"//div[@id='checkNumber{i}']//span[@class='icon icon-menu']")))
                                            action.move_to_element(actionsElement).click().perform()
                                            # actionsElement.click()
                                            time.sleep(2)
                                            multipleClaimsDownloadElement = WebDriverWait(driver, 60).until(
                                                EC.visibility_of_element_located((By.XPATH,
                                                                                  "//div[@class='dropdown-menu dropdown-menu-right show']//button[@role='menuitem'][contains(text(),'Download Check Summary and Multiple Claims Per Pag')]")))
                                            multipleClaimsDownloadElement.click()
                                            try:
                                                modalContentElement = WebDriverWait(driver, 10).until(
                                                    EC.visibility_of_all_elements_located(
                                                        (By.XPATH, "//div[@class='modal-content']")))
                                                continueElement = WebDriverWait(driver, 10).until(
                                                    EC.visibility_of_element_located(
                                                        (By.XPATH, "//button[normalize-space()='Continue']")))
                                                continueElement.click()
                                            except:
                                                pass
                                        else:
                                            continue
                                        time.sleep(8)
                                        downloadWaitCount = 0
                                        while downloadWaitCount < 120:
                                            list_of_files = glob.glob(f"{downloads_path}/*")
                                            try:
                                                latest_file = max(list_of_files, key=os.path.getmtime)
                                            except FileNotFoundError:
                                                time.sleep(1)
                                                downloadWaitCount += 1
                                                continue
                                            if latest_file != start_latest_file and latest_file.endswith('.pdf'):
                                                shutil.move(latest_file, sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf")
                                                break
                                            else:
                                                time.sleep(1)
                                                downloadWaitCount += 1
                                                continue
                                        if downloadWaitCount >= 120:
                                            print(
                                                f'The File with this check number --{checkNumber}--  is being downloaded for so long time please check')
                                            try:
                                                cancelElement = WebDriverWait(driver, 10).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, "//*[@class='icon icon-cancel cursor-pointer']")))
                                                cancelElement.click()
                                            except:
                                                pass
                                        if os.path.exists(sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf"):
                                            chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                            chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                            chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                            chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                            chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                            chequeWorksheet.cell(row=length + 2, column=6).value = "File Downloaded"
                                            try:
                                                chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                            except KeyError:
                                                chequeWorksheet.cell(row=length + 2,
                                                               column=7).value = 'Payee not found in Master'
                                            chequeWorkbook.save(chequeDetailsFilename)
                                            df = pd.read_excel(chequeDetailsFilename)
                                            df = df.replace(np.nan, '')
                                            length = len(df)
                                        else:
                                            chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                            chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                            chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                            chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                            chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                            chequeWorksheet.cell(row=length + 2, column=6).value = "File Not Downloaded"
                                            try:
                                                chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                            except KeyError:
                                                chequeWorksheet.cell(row=length + 2,
                                                               column=7).value = 'Payee not found in Master'
                                            chequeWorkbook.save(chequeDetailsFilename)
                                            df = pd.read_excel(chequeDetailsFilename)
                                            df = df.replace(np.nan, '')
                                            length = len(df)
                        else:
                            lastPageCheck = False
                            while endCount != totalCount or lastPageCheck == True:
                                for i in range(countPerPage):
                                    # print(i)
                                    rowElement = WebDriverWait(driver, 60).until(
                                        EC.visibility_of_all_elements_located((By.XPATH, f"//*[@id='checkNumber{i}']")))
                                    rowText = rowElement[0].text
                                    checkNumber = rowText.split('\n')[0]
                                    payer = re.sub(r"[:?/\|''*<>]", "", rowText.split('\n')[1])
                                    payee = re.sub(r"[:?/\|''*<>]", "", rowText.split('\n')[2])
                                    checkDate = rowText.split('\n')[3]
                                    receivedByAvailityDate = rowText.split('\n')[4]
                                    checkAmt = rowText.split('\n')[5]
                                    # print(checkNumber)
                                    # print(payer)
                                    # print(payee)
                                    # print(checkAmt)
                                    if receivedByAvailityDate == todayDate1 or receivedByAvailityDate == yesterday_formatted or checkDate == todayDate1 or checkDate == yesterday_formatted:
                                        try:
                                            if checkNumber not in uniqueCheckList:
                                                pdfIconElement = WebDriverWait(driver, 5).until(
                                                    EC.visibility_of_element_located((By.ID, f"checkeob{i}")))
                                                action.move_to_element(pdfIconElement).click().perform()
                                                # pdfIconElement.click()
                                            else:
                                                continue
                                            time.sleep(2)
                                            try:
                                                checkboxElement = WebDriverWait(driver, 10).until(
                                                    EC.visibility_of_element_located(
                                                        (By.XPATH, f"//label[@for='allcheckeob{i}']")))
                                                checkboxElement.click()
                                                time.sleep(2)
                                                downloadElement = WebDriverWait(driver, 10).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, "//button[normalize-space()='Download']")))
                                                downloadElement.click()
                                            except:
                                                pass
                                            time.sleep(8)
                                            downloadWaitCount = 0
                                            while downloadWaitCount < 120:
                                                list_of_files = glob.glob(f"{downloads_path}/*")
                                                try:
                                                    latest_file = max(list_of_files, key=os.path.getmtime)
                                                except FileNotFoundError:
                                                    time.sleep(1)
                                                    downloadWaitCount += 1
                                                    continue
                                                if latest_file != start_latest_file and latest_file.endswith('.pdf'):
                                                    shutil.move(latest_file, sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf")
                                                    break
                                                else:
                                                    time.sleep(1)
                                                    downloadWaitCount += 1
                                                    continue
                                            if downloadWaitCount >= 120:
                                                print(
                                                    f'The File with this check number --{checkNumber}--  is being downloaded for so long time please check')
                                                try:
                                                    closeElement = WebDriverWait(driver, 10).until(
                                                        EC.element_to_be_clickable(
                                                            (By.XPATH, "//button[normalize-space()='Close']")))
                                                    closeElement.click()
                                                except:
                                                    pass
                                                try:
                                                    cancelElement = WebDriverWait(driver, 10).until(
                                                        EC.element_to_be_clickable((By.XPATH,
                                                                                    "//*[@class='icon icon-cancel cursor-pointer']")))
                                                    cancelElement.click()
                                                except:
                                                    pass
                                            if os.path.exists(sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf"):
                                                chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                                chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                                chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                                chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                                chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                                chequeWorksheet.cell(row=length + 2, column=6).value = "File Downloaded"
                                                try:
                                                    chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                                except KeyError:
                                                    chequeWorksheet.cell(row=length + 2,
                                                                   column=7).value = 'Payee not found in Master'
                                                chequeWorkbook.save(chequeDetailsFilename)
                                                df = pd.read_excel(chequeDetailsFilename)
                                                df = df.replace(np.nan, '')
                                                length = len(df)
                                            else:
                                                chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                                chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                                chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                                chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                                chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                                chequeWorksheet.cell(row=length + 2, column=6).value = "File Not Downloaded"
                                                try:
                                                    chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                                except KeyError:
                                                    chequeWorksheet.cell(row=length + 2,
                                                                   column=7).value = 'Payee not found in Master'
                                                chequeWorkbook.save(chequeDetailsFilename)
                                                df = pd.read_excel(chequeDetailsFilename)
                                                df = df.replace(np.nan, '')
                                                length = len(df)
                                            try:
                                                closeElement = WebDriverWait(driver, 10).until(
                                                    EC.element_to_be_clickable(
                                                        (By.XPATH, "//button[normalize-space()='Close']")))
                                                closeElement.click()
                                            except:
                                                pass

                                        except:
                                            if checkNumber not in uniqueCheckList:
                                                actionsElement = WebDriverWait(driver, 60).until(
                                                    EC.visibility_of_element_located((By.XPATH,
                                                                                      f"//div[@id='checkNumber{i}']//span[@class='icon icon-menu']")))
                                                action.move_to_element(actionsElement).click().perform()
                                                # actionsElement.click()
                                                time.sleep(2)
                                                multipleClaimsDownloadElement = WebDriverWait(driver, 60).until(
                                                    EC.visibility_of_element_located((By.XPATH,
                                                                                      "//div[@class='dropdown-menu dropdown-menu-right show']//button[@role='menuitem'][contains(text(),'Download Check Summary and Multiple Claims Per Pag')]")))
                                                multipleClaimsDownloadElement.click()
                                                try:
                                                    modalContentElement = WebDriverWait(driver, 10).until(
                                                        EC.visibility_of_all_elements_located(
                                                            (By.XPATH, "//div[@class='modal-content']")))
                                                    continueElement = WebDriverWait(driver, 10).until(
                                                        EC.visibility_of_element_located(
                                                            (By.XPATH, "//button[normalize-space()='Continue']")))
                                                    continueElement.click()
                                                except:
                                                    pass
                                            else:
                                                continue
                                            time.sleep(8)
                                            downloadWaitCount = 0
                                            while downloadWaitCount < 120:
                                                list_of_files = glob.glob(f"{downloads_path}/*")
                                                try:
                                                    latest_file = max(list_of_files, key=os.path.getmtime)
                                                except FileNotFoundError:
                                                    time.sleep(1)
                                                    downloadWaitCount += 1
                                                    continue
                                                if latest_file != start_latest_file and latest_file.endswith('.pdf'):
                                                    shutil.move(latest_file, sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf")
                                                    break
                                                else:
                                                    time.sleep(1)
                                                    downloadWaitCount += 1
                                                    continue
                                            if downloadWaitCount >= 120:
                                                print(
                                                    f'The File with this check number --{checkNumber}--  is being downloaded for so long time please check')
                                                try:
                                                    cancelElement = WebDriverWait(driver, 10).until(
                                                        EC.element_to_be_clickable((By.XPATH,
                                                                                    "//*[@class='icon icon-cancel cursor-pointer']")))
                                                    cancelElement.click()
                                                except:
                                                    pass
                                            if os.path.exists(sharedParentPath + todayDate[0:2] + " " + current_month + "'" + todayDate[4:] + "\\" + todayDate + "\\Availity EOB\\" + checkNumber + "_" + payer + "_" + payee + "_" + checkAmt + ".pdf"):
                                                chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                                chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                                chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                                chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                                chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                                chequeWorksheet.cell(row=length + 2, column=6).value = "File Downloaded"
                                                try:
                                                    chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                                except KeyError:
                                                    chequeWorksheet.cell(row=length + 2,
                                                                   column=7).value = 'Payee not found in Master'
                                                chequeWorkbook.save(chequeDetailsFilename)
                                                df = pd.read_excel(chequeDetailsFilename)
                                                df = df.replace(np.nan, '')
                                                length = len(df)
                                            else:
                                                chequeWorksheet.cell(row=length + 2, column=1).value = checkNumber
                                                chequeWorksheet.cell(row=length + 2, column=2).value = payer
                                                chequeWorksheet.cell(row=length + 2, column=3).value = payee
                                                chequeWorksheet.cell(row=length + 2, column=4).value = checkDate
                                                chequeWorksheet.cell(row=length + 2, column=5).value = checkAmt
                                                chequeWorksheet.cell(row=length + 2, column=6).value = "File Not Downloaded"
                                                try:
                                                    chequeWorksheet.cell(row=length + 2, column=7).value = rsmName
                                                except KeyError:
                                                    chequeWorksheet.cell(row=length + 2,
                                                                   column=7).value = 'Payee not found in Master'
                                                chequeWorkbook.save(chequeDetailsFilename)
                                                df = pd.read_excel(chequeDetailsFilename)
                                                df = df.replace(np.nan, '')
                                                length = len(df)
                                if lastPageCheck:
                                    lastPageCheck = False
                                if endCount != totalCount:
                                    nextbuttonElement = WebDriverWait(driver, 60).until(
                                        EC.visibility_of_element_located(
                                            (By.XPATH, "(//button[normalize-space()='Next'])[1]")))
                                    nextbuttonElement.click()
                                    paginatorElement = WebDriverWait(driver, 60).until(EC.visibility_of_element_located(
                                        (By.XPATH, "//div[@class='align-items-center row']//span[@class='mx-2 mt-2']")))
                                    paginatorText = paginatorElement.text
                                    paginatorText = paginatorText.strip()
                                    # print(paginatorText)
                                    startCount = int(paginatorText.split(' ')[1])
                                    endCount = int(paginatorText.split(' ')[3])
                                    countPerPage = (endCount - startCount) + 1
                                    if 'more' not in paginatorText:
                                        totalCount = int(paginatorText.split(' ')[5])
                                    elif 'more' in paginatorText:
                                        totalCount = int(paginatorText.split(' ')[7])
                                    if endCount == totalCount:
                                        lastPageCheck = True
                elif 'no matches found' in stateNameFromPortal.strip().lower():
                    worksheet.cell(row=index + 2, column=6).value = "State Not found"
                    workbook.save(filename)
                    continue
                worksheet.cell(row=index + 2, column=6).value = "Processed"
                workbook.save(filename)