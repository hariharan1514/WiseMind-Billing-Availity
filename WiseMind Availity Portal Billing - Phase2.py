import os.path
from dataclasses import replace
from datetime import datetime
import pandas as pd
import sys

from pymsgbox import confirm
from selenium.common import NoSuchElementException
from selenium.webdriver import ActionChains

parent_folder_path = r"Z:\Wisemind\Charge Entry -Billing\Billing Dates"
path_temp_date = datetime.today().strftime('%m%d%Y')
bcbs_file_path = (parent_folder_path + "\\" +path_temp_date[4:] + "\\" +datetime.today().strftime("%m %b'%Y") +
                   "\\" +path_temp_date + "\\" +f"BCBS scrubbed file - {path_temp_date}.xlsx")

### Check whether the Straightforward Billing file is present or not.
if not os.path.exists(bcbs_file_path):
    print("Run the 'WiseMind Availity Billing Phase1.py' script first, then execute the Phase2 script.")
    sys.exit()
else:
    bcbs_billing_df = pd.read_excel(bcbs_file_path, sheet_name=0)
    print(f"Staright Forward Data Row Count: {len(bcbs_billing_df)}")
    ### Check if the Straightforward Billing file contains at least one data row. ###
    if bcbs_billing_df.shape[0] == 0:
        print("No BCBS Claims detected in today's run. Script completed successfully with no records to process. \n\nExiting gracefully as per expected behavior.")
        sys.exit()
    else:
        import time
        from openpyxl import load_workbook
        from selenium import webdriver
        from selenium.webdriver.ie.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.wait import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import StaleElementReferenceException
        from selenium.webdriver.common.keys import Keys
        import re
        from openpyxl import Workbook, load_workbook

        # Read Configuration Sheet with openpyxl & Pandas module
        config_sheet_path = r"Z:\Wisemind\Charge Entry -Billing\Automation Config File\ConfigSheet.xlsx"
        # Get the Username & Password
        master_workbook = load_workbook(config_sheet_path, data_only=True)
        password_sheet = master_workbook[master_workbook.sheetnames[0]]
        username = password_sheet['B1'].value
        password = password_sheet['B2'].value

        # Initiate the Chrome instance
        chrome_option = webdriver.ChromeOptions()
        chrome_option.add_experimental_option('detach', True)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_option)
        driver.maximize_window()
        actions = ActionChains(driver)

        driver.get("https://app.theranest.com/login") # Launching the router

        # send the username & password to the represented field
        try:
            username_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='Email']")))
            username_element.send_keys(username)
        except:
            print("Login issue, Please login after some time.!")
            sys.exit()

        password_element = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='Password']")))
        password_element.send_keys(password)

        login_button_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//button[normalize-space()='Log In']")))
        login_button_element.click()

        try:
            mainpage_element = WebDriverWait(driver,90).until(EC.visibility_of_all_elements_located((By.XPATH,"//div[@role='group']")))
        except:
            login_button_element = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Log In']")))
            print("Please try logging in again after some time. The site is currently experiencing login issues.")
            driver.close()
            sys.exit()

        bcbs_billing_wb = load_workbook(bcbs_file_path, data_only=True)
        bcbs_sheet = bcbs_billing_wb.active

        ### Add Exception Column if not present
        exception_col_name = ["Active/Archived", "Transaction Status"]
        headers = [cell.value for cell in bcbs_sheet[1]]
        next_col_index = len(headers) + 1
        for col in exception_col_name:
            if col not in headers:
                bcbs_sheet.cell(row=1,column=next_col_index,value=col)
                next_col_index += 1

        bcbs_billing_wb.save(bcbs_file_path)

        ### Build a dictionary mapping header names to column indices
        data_columns = {}
        for col in range(1, bcbs_sheet.max_column + 1):
            col_name = bcbs_sheet.cell(row=1, column=col).value
            if col_name:
                data_columns[col_name.strip()] = col
        for row in range(2, bcbs_sheet.max_row + 1):

            status = bcbs_sheet.cell(row=row, column=data_columns['Status']).value
            transaction_status = bcbs_sheet.cell(row=row, column=data_columns['Transaction Status']).value
            # exception = bcbs_sheet.cell(row=row, column=data_columns['Exceptions']).value
            if status != "Yes" and transaction_status == "Transaction Number updated":
                continue

            client_name = bcbs_sheet.cell(row=row, column=data_columns['Client Name']).value
            client_id_number = bcbs_sheet.cell(row=row, column=data_columns['Client ID Number']).value
            invoice_number = bcbs_sheet.cell(row=row, column=data_columns['Claim Invoice Number']).value
            transaction_number = bcbs_sheet.cell(row=row, column=data_columns['Transaction Number']).value
            claimTab_status = bcbs_sheet.cell(row=row, column=data_columns['Active/Archived']).value


            print(f"Patient Name: {client_name}")
            partial_name = client_name.split()[0]
            driver.get("https://wisemind71.theranest.com/clients")
            time.sleep(3)

            tabs = ["Active","Archived"]
            name_to_search = [client_name,partial_name]

            found_data = False
            active_tab = False
            archived_tab =False

            for tab in tabs:
                for name in name_to_search:

                    for attempt in range(3):
                        try:
                            tab_element = WebDriverWait(driver,120).until(
                                EC.presence_of_element_located((By.XPATH,f"(//span[contains(text(), '{tab}')])[1]")))
                            time.sleep(1)
                            tab_element.click()
                            time.sleep(2)
                            break
                        except StaleElementReferenceException:
                            continue

                    search_bar_element = WebDriverWait(driver, 120).until(
                        EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
                    search_bar_element.clear()
                    search_bar_element.send_keys(name)
                    time.sleep(2)
                    #Check if data exists
                    try:
                        # driver.find_element(By.XPATH,"//div[text()='No data available']")
                        no_data_element = WebDriverWait(driver,5).until(EC.visibility_of_element_located((By.XPATH,"//div[text()='No data available']")))
                        search_bar_element.clear()

                    except:
                        found_data = True
                        clientname_table_element = WebDriverWait(driver,120).until(EC.visibility_of_all_elements_located((By.XPATH,"//table[contains(@class, 'k-grid-table')]//tr")))
                        for tbl_row in range(1,len(clientname_table_element) +1):
                            time.sleep(2)
                            if tab == 'Active':
                                client_name_element =WebDriverWait(driver,120).until(EC.element_to_be_clickable((By.XPATH,f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                client_name_element.click()

                                ### Check Client ID found or not
                                client_id_element_number = None
                                try:
                                    client_id_element = WebDriverWait(driver,30).until(EC.visibility_of_element_located((By.XPATH,"//div[@data-aqa='ClientID']//div[2]")))
                                    client_id_element_number = client_id_element.text

                                except:
                                    driver.back()
                                    found_data = False

                                if client_id_number == client_id_element_number:
                                    found_data = True
                                    active_tab = True

                                    claim_url = driver.current_url
                                    firstname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='FirstName']"))).get_attribute("value")
                                    middlename = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='MiddleName']"))).get_attribute("value")
                                    lastname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='LastName']"))).get_attribute("value")

                                    portal_client_name = f"{firstname} {middlename} {lastname}".strip()
                                    portal_client_name = " ".join(portal_client_name.split())
                                    break
                                else:
                                    found_data =False
                                    driver.back()
                            elif tab == 'Archived':
                                archived_client_id_element = WebDriverWait(driver,120).until(EC.visibility_of_element_located((By.XPATH,f"//table//tbody//tr[{tbl_row}]/td[3]"))) #//table//tbody//tr[1]/td[3]
                                archived_client_id_number = archived_client_id_element.text
                                if client_id_number == archived_client_id_number:
                                    archived_tab = True
                                    unarchive_button_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,f"(//button[@data-aqa='btnUnarchive'])[{tbl_row}]")))
                                    unarchive_button_element.click()

                                    submit_button_element = WebDriverWait(driver,120).until(EC.element_to_be_clickable((By.XPATH,'//button[@data-aqa="btnSubmit"]')))
                                    submit_button_element.click()
                                    time.sleep(1)

                                    for attempt in range(3):
                                        try:
                                            tab_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), 'Active')])[1]")))
                                            time.sleep(1)
                                            tab_element.click()
                                            time.sleep(2)
                                            break
                                        except StaleElementReferenceException:
                                            continue

                                    for name in name_to_search:
                                        search_bar_element = WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
                                        search_bar_element.clear()
                                        search_bar_element.send_keys(name)
                                        time.sleep(2)
                                        try:
                                            # driver.find_element(By.XPATH, "//div[text()='No data available']")
                                            no_data_element = WebDriverWait(driver,120).until(EC.visibility_of_element_located((By.XPATH,"//div[text()='No data available']")))
                                            search_bar_element.clear()
                                        except:
                                            found_data = True
                                            clientname_table_element = WebDriverWait(driver, 120).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[contains(@class, 'k-grid-table')]//tr")))
                                            for tbl_row in range(1, len(clientname_table_element) + 1):
                                                time.sleep(2)
                                                for attempt in range(3):
                                                    try:
                                                        client_name_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                                        time.sleep(1)
                                                        client_name_element.click()
                                                        break
                                                    except StaleElementReferenceException:
                                                        continue
                                                client_id_element = WebDriverWait(driver, 40).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-aqa='ClientID']//div[2]")))
                                                client_id_element_number = client_id_element.text

                                                if client_id_number != client_id_element_number:
                                                    found_data = False
                                                    archived_tab = False
                                                    driver.back()
                                                else:
                                                    claim_url = driver.current_url
                                                    firstname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='FirstName']"))).get_attribute("value")
                                                    middlename = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='MiddleName']"))).get_attribute("value")
                                                    lastname = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='LastName']"))).get_attribute("value")

                                                    portal_client_name = f"{firstname} {middlename} {lastname}".strip()
                                                    portal_client_name = " ".join(portal_client_name.split())
                                                    break
                                        if archived_tab:
                                            break
                            if found_data:
                                break
                        if archived_tab or found_data:
                            break
                if found_data:
                    break

            if not found_data:
                continue

            if active_tab or claimTab_status != "Archived":
                bcbs_sheet.cell(row=row, column=data_columns['Active/Archived']).value = "Active"
                bcbs_billing_wb.save(bcbs_file_path)

            elif archived_tab:
                bcbs_sheet.cell(row=row, column=data_columns['Active/Archived']).value = "Archived"
                bcbs_billing_wb.save(bcbs_file_path)


            ### Navigate to ledger ###
            ledger_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"//a[@aria-label='Ledger']")))
            ledger_btn_element.click()

            openinvoice_btn_element = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH,"//a[@data-aqa='openInvoices']")))
            openinvoice_btn_element.click()

            # recent_btn = WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnRecentDays']")))
            # recent_btn.click()

            time.sleep(2)

            ### Services Table looping ###
            dos_match = False
            # extracted_dos_date = dos_date.replace(" ET", "").strip()
            service_table_element = WebDriverWait(driver, 120).until(EC.presence_of_all_elements_located((By.XPATH,"(//table[contains(@class, 'k-grid-table')]//tr)")))
            driver.execute_script("document.body.style.zoom= '50%'")
            # if len(service_table_element) >= 1 :
            for tbl_row in range(1, len(service_table_element) +1):

                service_tblrow_element = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH,f"//table[@class='k-grid-table']//tr[{tbl_row}]//td[@data-aqa='number']")))
                extracted_invoice_number = service_tblrow_element.text

                if invoice_number == extracted_invoice_number :#and service_type == title_value
                    # dos_match = True
                    print(f"Invoice Number: ({invoice_number}) matched.")
                    service_tblrow_element.click()
                    invoice_footnote_element = WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH,"//textarea[@data-aqa='invoiceFootnot']")))
                    invoice_footnote_element.click()
                    invoice_footnote_element.send_keys(f"Claim submitted through availity portal under transaction ID#{invoice_number}")
                    save_button_element = WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH,"(//button[@data-aqa='saveInvoic'])[2]")))
                    bcbs_sheet.cell(row=row, column=data_columns['Transaction Status']).value = "Transaction Number updated"
                    bcbs_billing_wb.save(bcbs_file_path)

                    time.sleep(2)

            if archived_tab or claimTab_status == "Archived":
                driver.get("https://wisemind71.theranest.com/clients")
            for attempt in range(3):
                try:
                    tab_element = WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), 'Active')])[1]")))
                    tab_element.click()
                    time.sleep(2)
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
                    continue
            search_bar_element = WebDriverWait(driver, 120).until(
                EC.visibility_of_element_located((By.XPATH, '//div[@data-aqa="inputFullName"]//input')))
            search_bar_element.clear()
            search_bar_element.send_keys(portal_client_name)
            time.sleep(1)
            try:
                no_data_element = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.XPATH, "//div[text()='No data available']")))
                search_bar_element.clear()
            except:
                clientname_table_element = WebDriverWait(driver, 120).until(
                    EC.visibility_of_all_elements_located(
                        (By.XPATH, "//table[contains(@class, 'k-grid-table')]//tr")))
                if len(clientname_table_element) > 1:
                    for tbl_row in range(1, len(clientname_table_element) + 1):
                        time.sleep(2)
                        for attempt in range(3):
                            try:
                                client_name_element = WebDriverWait(driver, 120).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, f"//table//tbody//tr[{tbl_row}]/td[2]/span/a")))
                                time.sleep(1)
                                client_name_element.click()
                                break
                            except StaleElementReferenceException:
                                continue
                        client_id_element = WebDriverWait(driver, 40).until(
                            EC.visibility_of_element_located((By.XPATH, "//div[@data-aqa='ClientID']//div[2]")))
                        client_id_element_number = client_id_element.text

                        if client_id_number != client_id_element_number:
                            driver.back()
                        else:
                            driver.back()
                            archived_btn_element = WebDriverWait(driver, 30).until(EC.element_to_be_clickable(
                                (By.XPATH, f"(//button[@data-aqa='btnArchive'])[{tbl_row}]")))
                            archived_btn_element.click()

                            arcSubmit_element = WebDriverWait(driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnSubmit']")))
                            arcSubmit_element.click()
                            break
                else:
                    archived_btn_element = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "(//button[@data-aqa='btnArchive'])[1]")))
                    archived_btn_element.click()

                    arcSubmit_element = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[@data-aqa='btnSubmit']")))
                    arcSubmit_element.click()

        time.sleep(2)
        driver.get("https://wisemind71.theranest.com/home/logout")
        time.sleep(3)

        print("✅✅ Billing Completed! ✅✅")
        driver.quit()